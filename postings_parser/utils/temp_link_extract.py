import openpyxl

# Load the workbook and select the sheet
workbook = openpyxl.load_workbook('HiringTechCompanies.xlsx')
sheet = workbook['TheList']  # Change 'Sheet1' to your sheet name

# Specify the column from which to extract hyperlinks
column_letter = 'A'  # Change 'A' to the column letter you need

hyperlinks = []

# Iterate through the cells in the specified column
for row in sheet.iter_rows(min_col=sheet[column_letter + '1'].col_idx, 
                           max_col=sheet[column_letter + '1'].col_idx,
                           min_row=2,  # Assuming the first row is the header
                           max_row=sheet.max_row):
    for cell in row:
        if cell.hyperlink:
            hyperlinks.append(cell.hyperlink.target+"\n")

# Print the extracted hyperlinks
with open("tech_company_urls.txt", "w") as f:
    f.writelines(hyperlinks)
